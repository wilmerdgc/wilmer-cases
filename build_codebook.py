from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
SECTION_FONT = Font(bold=True, size=12, color="1F4E78")
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = BORDER


def style_body(ws, start_row, end_row, num_cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_section(ws, row, title, span):
    ws.cell(row=row, column=1, value=title)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[row].height = 22


def write_table(ws, headers, rows, start_row=1):
    for i, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=i, value=h)
    style_header(ws, start_row, len(headers))
    ws.row_dimensions[start_row].height = 24
    for r_i, row in enumerate(rows, start=start_row + 1):
        for c_i, val in enumerate(row, start=1):
            ws.cell(row=r_i, column=c_i, value=val)
    if rows:
        style_body(ws, start_row + 1, start_row + len(rows), len(headers))
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


# =============================================================
# Sheet: Overview
# =============================================================
ws = wb.create_sheet("Overview")
set_widths(ws, [28, 80])
ws["A1"] = "Wilmer Cases — Field Codebook"
ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
ws.merge_cells("A1:B1")
ws["A2"] = "Generated for index.html (single-file web app) — all data fields, enums, and storage keys."
ws.merge_cells("A2:B2")
ws["A2"].font = Font(italic=True, color="595959")

overview_rows = [
    ("Sheet", "Contents"),
    ("Record (events)", "Fields on each entry in records[] — hearings, tasks, deadlines, reminders"),
    ("Record.history", "Action-log entries attached to a record"),
    ("CaseMeta — Client", "Client/respondent demographics in caseMeta[clid]"),
    ("CaseMeta — Stages", "Procedural stage fields (FC / IC / USCIS)"),
    ("CaseMeta — Docs & Notes", "Document sections, case notes, I-360 notice"),
    ("CaseMeta — Respondents", "Co-respondents / riders"),
    ("CaseMeta — Issues", "Legal-issue tracker (caseIssues)"),
    ("Firm Info", "firmInfo object — attorney/firm data used for G-28"),
    ("Settings Arrays", "paralegals, eventTypes, caseTypes, venues, customColumns, etc."),
    ("Judges", "Entries in judges[]"),
    ("Event Type Meta", "eventTypeMeta — category & icon per event type"),
    ("G-28 PDF Mapping", "PDF field names mapped to firmInfo / caseMeta values"),
    ("Enumerations", "Valid values for enum-style fields"),
    ("Storage Keys", "Firestore collections and LocalStorage keys"),
]
write_table(ws, overview_rows[0], overview_rows[1:], start_row=4)

# =============================================================
# Sheet: Record (events)
# =============================================================
ws = wb.create_sheet("Record (events)")
set_widths(ws, [26, 14, 50, 40])
headers = ["Field", "Type", "Description", "Values / Example"]
rows = [
    ("id", "string", "Unique record identifier", '"22738", "r17737797826008nu3"'),
    ("clid", "string", "Case ID — links multiple records to one case", '"22738"'),
    ("client", "string", "Client / respondent name", '"Marielys Marquez"'),
    ("anum", "string", "USCIS A-Number", '"220-347-638", "N/A"'),
    ("type", "string (enum)", "Case type — must match caseTypes[]", '"Defensive Asylum"'),
    ("event", "string (enum)", "Event type — must match eventTypes[]", '"MASTER", "TASK"'),
    ("date", "string", "Event date", "YYYY-MM-DD e.g. 2026-03-11"),
    ("timeStart", "string", "Start time", "HH:MM e.g. 09:00"),
    ("timeEnd", "string", "End time", "HH:MM e.g. 15:00"),
    ("atty", "string", "Attorney / paralegal — must match paralegals[]", '"Jessica Vicuna"'),
    ("assigned", "string", "Assignee", '"Me" or paralegal name'),
    ("notes", "string", "Free-text notes (multiline)", '"Motion to Dispense Service"'),
    ("finishedStatus", "string (enum)", "Completion marker", '"completed", "attended", "cancelled", "rescheduled", null'),
    ("deletedStatus", "string", "Soft-delete flag", '"deleted" or null'),
    ("deletedDate", "string", "When soft-deleted", "YYYY-MM-DD"),
    ("checklist", "array", "Inline checklist items for this event", '["File motion", ...]'),
    ("docs", "array", "Attached documents", '[{name:"file.pdf"}, ...]'),
    ("history", "array", "Action log (see Record.history sheet)", "Array of history entries"),
    ("hearingMode", "string (enum)", "Hearing format", '"in-person", "virtual", "webex", "internet-based"'),
    ("soIssueDate", "string", "Scheduling Order issue date", "YYYY-MM-DD"),
    ("_parentId", "string", "Parent event (used by ⏰ reminders)", "Record id of parent"),
    ("custom_*", "varies", "User-defined custom column key", "custom_address_1775582040709"),
]
write_table(ws, headers, rows)

# =============================================================
# Sheet: Record.history
# =============================================================
ws = wb.create_sheet("Record.history")
set_widths(ws, [18, 14, 50, 30])
headers = ["Field", "Type", "Description", "Example"]
rows = [
    ("type", "string (enum)", "Action performed", '"completed", "attended", "cancelled", "rescheduled", "deleted", "note"'),
    ("date", "string", "When the action occurred", '"Apr 22, 2026"'),
    ("note", "string", "Optional description", '"Marked done from week view"'),
    ("prevDate", "string", "Previous date (rescheduled only)", '"2026-03-18"'),
    ("newDate", "string", "New date (rescheduled only)", '"2026-06-16"'),
]
write_table(ws, headers, rows)

# =============================================================
# Sheet: CaseMeta — Client
# =============================================================
ws = wb.create_sheet("CaseMeta — Client")
set_widths(ws, [22, 14, 50])
headers = ["Field", "Type", "Description"]
rows = [
    ("clientFirstName", "string", "Given name"),
    ("clientMiddleName", "string", "Middle name"),
    ("clientLastName", "string", "Family name"),
    ("clientEmail", "string", "Email"),
    ("clientPhone", "string", "Phone"),
    ("clientMobile", "string", "Mobile"),
    ("clientStreet", "string", "Street address"),
    ("clientApt", "string", "Apt / Suite / Floor"),
    ("clientCity", "string", "City"),
    ("clientState", "string", "State"),
    ("clientZip", "string", "ZIP code"),
    ("clientCountry", "string", "Country"),
]
write_table(ws, headers, rows)

# =============================================================
# Sheet: CaseMeta — Stages
# =============================================================
ws = wb.create_sheet("CaseMeta — Stages")
set_widths(ws, [18, 16, 40, 40])
headers = ["Field", "Type", "Description", "Values"]
rows = [
    ("fcStage", "string (enum)", "Family Court stage", '"initial", "petition", "trial", "post-trial"'),
    ("icStage", "string (enum)", "Immigration Court stage", '"no-info", "master-prep", "master", "individual"'),
    ("icStageDate", "string", "Date case entered the current IC stage", "YYYY-MM-DD"),
    ("uscisStage", "string", "USCIS stage (partial implementation)", "—"),
    ("consolidated", "boolean", "Case is consolidated with others", "true / false"),
    ("leadIdx", "number", "Index of lead in respondents[]", "0"),
]
write_table(ws, headers, rows)

# =============================================================
# Sheet: CaseMeta — Docs & Notes
# =============================================================
ws = wb.create_sheet("CaseMeta — Docs & Notes")
set_widths(ws, [18, 14, 70])
headers = ["Field", "Type", "Description / Structure"]
rows = [
    ("docSections", "array", "Sectioned document sets: [{sectionName, docs:[{name}]}, …]"),
    ("caseNotesList", "array", "Timestamped notes: [{date, text}, …]"),
    ("caseNotes", "string", "Legacy single notes field"),
    ("caseNotes2", "string", "Legacy secondary notes field"),
    ("i360Notice", "object", "I-360 approval notice: {name, uploadedAt}"),
    ("i360Notice.name", "string", "Filename"),
    ("i360Notice.uploadedAt", "string", "Upload date (YYYY-MM-DD)"),
]
write_table(ws, headers, rows)

# =============================================================
# Sheet: CaseMeta — Respondents
# =============================================================
ws = wb.create_sheet("CaseMeta — Respondents")
set_widths(ws, [16, 14, 50, 30])
headers = ["Field", "Type", "Description", "Example / Values"]
rows = [
    ("name", "string", "Full name", '"Carlos Reyes"'),
    ("firstName", "string", "Given name", '"Carlos"'),
    ("middleName", "string", "Middle name", '"Miguel"'),
    ("lastName", "string", "Family name", '"Reyes"'),
    ("anum", "string", "A-Number", '"246-381-258"'),
    ("relationship", "string", "Relationship to lead", '"sibling", "spouse", "child"'),
    ("isLead", "boolean", "Lead respondent flag", "true / false"),
    ("atty", "string", "Assigned attorney", '"Jessica Vicuna"'),
    ("judgeIC", "string", "References judges[].id for IC", '"j1"'),
    ("judgeFC", "string", "References judges[].id for FC", '"j22"'),
]
write_table(ws, headers, rows)

# =============================================================
# Sheet: CaseMeta — Issues
# =============================================================
ws = wb.create_sheet("CaseMeta — Issues")
set_widths(ws, [22, 14, 60])
headers = ["Field", "Type", "Description"]
rows = [
    ("caseIssues", "object", "Keyed by caseIssueTypes[].id; value is an issue entry (below)"),
    ("caseIssues[issueId].checked", "boolean", "Issue applies to this case"),
    ("caseIssues[issueId].note", "string", "Notes about the issue"),
    ("judgeIC", "string", "Immigration Court judge ID (references judges[].id)"),
    ("judgeFC", "string", "Family Court judge/referee ID (references judges[].id)"),
]
write_table(ws, headers, rows)

# =============================================================
# Sheet: Firm Info
# =============================================================
ws = wb.create_sheet("Firm Info")
set_widths(ws, [22, 14, 10, 50])
headers = ["Field", "Type", "G-28 Part", "Description"]
rows = [
    ("lastName", "string", "Pt1", "Attorney family name"),
    ("firstName", "string", "Pt1", "Attorney given name"),
    ("middleName", "string", "Pt1", "Attorney middle name"),
    ("firmName", "string", "Pt2", "Law firm name"),
    ("barNumber", "string", "Pt1", "Bar number"),
    ("licensingAuthority", "string", "Pt1", "States licensed in"),
    ("street", "string", "Pt1", "Firm street"),
    ("apt", "string", "Pt1", "Suite / floor"),
    ("city", "string", "Pt1", "Firm city"),
    ("state", "string", "Pt1", "Firm state"),
    ("zip", "string", "Pt1", "Firm ZIP"),
    ("country", "string", "Pt1", "Firm country"),
    ("phone", "string", "Pt1", "Daytime phone"),
    ("mobile", "string", "Pt1", "Mobile"),
    ("email", "string", "Pt1", "Email"),
    ("fax", "string", "Pt1", "Fax"),
]
write_table(ws, headers, rows)

# =============================================================
# Sheet: Settings Arrays
# =============================================================
ws = wb.create_sheet("Settings Arrays")
set_widths(ws, [24, 16, 70])
headers = ["Name", "Type", "Contents"]
rows = [
    ("paralegals", "string[]", "Attorney / paralegal names (drives atty and assigned fields)"),
    ("eventTypes", "string[]", "Allowed values for record.event"),
    ("caseTypes", "string[]", "Allowed values for record.type"),
    ("venues", "string[]", "Court locations"),
    ("callUpTypes", "string[]", "Events that trigger 30-day call-ups (subset of eventTypes)"),
    ("caseIssueTypes", "object[]", "{id, name, proceedings[]} — legal issues per case type"),
    ("customChecklists", "object[]", "{id, name, caseTypes[]}"),
    ("customColumns", "object[]", '{key, label, type:"text"|"select", options[]}'),
    ("colVis", "object", "Per-column visibility flags (keyed by column name)"),
    ("judges", "object[]", "See Judges sheet"),
    ("paralegalEmails", "object", "Map: paralegal name → email address"),
    ("eventTypeMeta", "object", "Map: event name → {category, icon}"),
    ("caseTypeDocLists", "object", "Map: case type → [{sectionName, docs:[{name}]}, …]"),
    ("caseTypeStages", "object", "Map: case type → [{name, events[], docs[], checklist[]}]"),
]
write_table(ws, headers, rows)

# =============================================================
# Sheet: Judges
# =============================================================
ws = wb.create_sheet("Judges")
set_widths(ws, [14, 14, 50, 30])
headers = ["Field", "Type", "Description", "Example"]
rows = [
    ("id", "string", "Unique judge identifier", '"j1"'),
    ("name", "string", "Judge's full name", '"Carrie Johnson-Papillo"'),
    ("venue", "string", "Court venue", '"26 FEDERAL PLAZA"'),
    ("role", "string", "Role", '"Judge" or "Referee"'),
    ("email", "string", "Email", "carrie.c.johnson-papillo@usdoj.gov"),
    ("clerk", "string", "Clerk phone", '"212-602-6603"'),
    ("webex", "string", "Accepts WebEx", '"yes" / "no"'),
    ("master", "string", "Master hearing rules (multiline)", "Procedural text"),
    ("individual", "string", "Individual hearing rules", "Procedural text"),
    ("bond", "string", "Bond / detention rules", "Procedural text"),
    ("motion", "string", "Motion preferences", "Procedural text"),
    ("notes", "string", "Freeform notes", "—"),
]
write_table(ws, headers, rows)

# =============================================================
# Sheet: Event Type Meta
# =============================================================
ws = wb.create_sheet("Event Type Meta")
set_widths(ws, [16, 14, 40, 40])
headers = ["Field", "Type", "Description", "Values"]
rows = [
    ("category", "string", "Used for row coloring & filtering", '"deadline", "hearing", "task", "administrative"'),
    ("icon", "string", "Display icon (emoji or text)", '"📅", "⚖️"'),
]
write_table(ws, headers, rows)

# =============================================================
# Sheet: G-28 PDF Mapping
# =============================================================
ws = wb.create_sheet("G-28 PDF Mapping")
set_widths(ws, [60, 28, 40])
headers = ["PDF Field Name", "Source", "Description"]
rows = [
    ("form1[0].#subform[0].Pt1Line2a_FamilyName[0]", "firmInfo.lastName", "Attorney last name"),
    ("form1[0].#subform[0].Pt1Line2b_GivenName[0]", "firmInfo.firstName", "Attorney first name"),
    ("form1[0].#subform[0].Pt1Line2c_MiddleName[0]", "firmInfo.middleName", "Attorney middle name"),
    ("form1[0].#subform[0].Pt1Line3a_StreetNumberName[0]", "firmInfo.street", "Street address"),
    ("form1[0].#subform[0].Pt1Line3c_CityOrTown[0]", "firmInfo.city", "City"),
    ("form1[0].#subform[0].Pt1Line3e_ZipCode[0]", "firmInfo.zip", "ZIP code"),
    ("form1[0].#subform[0].Pt1Line3h_Country[0]", "firmInfo.country", "Country"),
    ("form1[0].#subform[0].Pt1Line4_DaytimePhoneNumber1[0]", "firmInfo.phone", "Phone number"),
    ("form1[0].#subform[0].Pt1Line5_MobileNumber[0]", "firmInfo.mobile", "Mobile number"),
    ("form1[0].#subform[0].Pt1Line6_Email[0]", "firmInfo.email", "Email"),
    ("form1[0].#subform[0].Pt1Line7_FaxNumber[0]", "firmInfo.fax", "Fax number"),
    ("form1[0].#subform[0].Pt2Line1d_LawFirmName[0]", "firmInfo.firmName", "Law firm name"),
    ("form1[0].#subform[1].Pt3Line6a_FamilyName[0]", "caseMeta.clientLastName", "Client last name"),
    ("form1[0].#subform[1].Pt3Line6b_GivenName[0]", "caseMeta.clientFirstName", "Client first name"),
    ("form1[0].#subform[1].Pt3Line6c_MiddleName[0]", "caseMeta.clientMiddleName", "Client middle name"),
    ("form1[0].#subform[1].Pt3Line10_DaytimePhoneNumber1[0]", "caseMeta.clientPhone", "Client phone"),
    ("form1[0].#subform[1].Pt3Line11_MobileNumber[0]", "caseMeta.clientMobile", "Client mobile"),
    ("form1[0].#subform[1].Pt3Line12_Email[0]", "caseMeta.clientEmail", "Client email"),
    ("form1[0].#subform[1].Pt3Line13a_StreetNumberName[0]", "caseMeta.clientStreet", "Client street"),
    ("form1[0].#subform[1].Pt3Line13c_CityOrTown[0]", "caseMeta.clientCity", "Client city"),
    ("form1[0].#subform[1].Pt3Line13e_ZipCode[0]", "caseMeta.clientZip", "Client ZIP"),
    ("form1[0].#subform[1].Pt3Line13h_Country[0]", "caseMeta.clientCountry", "Client country"),
]
write_table(ws, headers, rows)

# =============================================================
# Sheet: Enumerations
# =============================================================
ws = wb.create_sheet("Enumerations")
set_widths(ws, [22, 80])
headers = ["Enum", "Valid Values"]
rows = [
    ("Case types", "Defensive Asylum | Affirmative Asylum | Special Juvenile in Court | Cancellation of Removal | Withholding of Removal | Voluntary Departure"),
    ("Event types", "MASTER | INDIVIDUAL | SCHEDULING ORDER | TASK | TASK FOR FC | TASK FOR 3DD | PREP MASTER | PREP USCIS INTERVIEW | CALL UP DEADLINE | 30-DAY DEADLINE | 60-DAY DEADLINE | 90-DAY DEADLINE | PRIORITY DAY | ICE Appointment | Initial Petition FC | FAMILY COURT | ⏰ Reminder"),
    ("Hearing modes", "in-person | virtual | webex | internet-based"),
    ("Finished status", "completed | attended | cancelled | rescheduled | null (active)"),
    ("FC stages", "initial | petition | trial | post-trial"),
    ("IC stages", "no-info | master-prep | master | individual"),
    ("History types", "completed | attended | cancelled | rescheduled | deleted | note"),
    ("Event categories", "deadline | hearing | task | administrative"),
    ("Judge roles", "Judge | Referee"),
    ("Custom column types", "text | select"),
]
write_table(ws, headers, rows)

# =============================================================
# Sheet: Storage Keys
# =============================================================
ws = wb.create_sheet("Storage Keys")
set_widths(ws, [20, 36, 50])
headers = ["Layer", "Key / Collection", "Purpose"]
rows = [
    ("Firestore", "wilmer/data", "Main app state (single doc); auth required"),
    ("Firestore", "wilmer_daynotes/{YYYY-MM-DD}", "Per-day priorities & notes; auth required"),
    ("LocalStorage", "casetracker_v1", "Main app JSON state"),
    ("LocalStorage", "wilmer_device_id", "Unique device identifier"),
    ("LocalStorage", "wilmer_save_epoch", "Last save timestamp (ms)"),
    ("LocalStorage", "last_autobackup_epoch", "Last auto-backup timestamp"),
    ("LocalStorage", "last_backup_label", "Last backup label"),
    ("LocalStorage", "defaultView", 'Default view on load ("today", "week", …)'),
    ("LocalStorage", "weekShowWeekend", '"0" or "1"'),
    ("LocalStorage", "g28-template", "Cached G-28 PDF template (base64)"),
    ("LocalStorage", "casetracker_venues", "Venues JSON"),
    ("LocalStorage", "casetracker_paralegal_emails", "Paralegal → email map"),
    ("LocalStorage", "casetracker_issue_types_v1", "Case issue types JSON"),
    ("LocalStorage", "casetracker_custom_checklists_v1", "Custom checklists JSON"),
    ("LocalStorage", "wilmer_judges_v1", "Judges JSON"),
    ("LocalStorage", "sij-col-widths", "Column widths JSON"),
    ("LocalStorage", "casetracker_versions_v1", "Backup version list"),
    ("LocalStorage", "migration_clientFirstName_v1", 'Migration flag ("done")'),
    ("LocalStorage", "migration_resync_record_client_v1", 'Migration flag ("done")'),
    ("LocalStorage", "{YYYY-MM-DD}", "Per-day note cache (keyed by date)"),
]
write_table(ws, headers, rows)

out_path = "/home/user/wilmer-cases/Wilmer_Cases_Codebook.xlsx"
wb.save(out_path)
print(f"Wrote {out_path}")
